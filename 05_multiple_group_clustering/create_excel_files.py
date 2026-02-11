#!/usr/bin/env python3
"""
Script to create Excel files for each iteration with comparison data.
Each Excel file will have:
- Tab 1: Metadata
- Tab 2: knockdown_PmSTAT_vs_knockdown_P (C2 only)
- Tab 3: WSSV_infection_vs_knockdown_PmS (C2 only)
- Tab 4: WSSV_infection_vs_knockdown_P-1 (C2 only)
"""

import os
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def read_metadata(metadata_file):
    """Read metadata from file and return as list of rows."""
    if not os.path.exists(metadata_file):
        return []
    
    with open(metadata_file, 'r') as f:
        lines = f.readlines()
    
    # Parse metadata - tab-separated values
    metadata_rows = []
    for line in lines:
        line = line.strip()
        if line:
            parts = line.split('\t')
            metadata_rows.append(parts)
    
    return metadata_rows

def read_two_column_file(file_path):
    """Read a two-column comparison file and return data."""
    if not os.path.exists(file_path):
        return []
    
    with open(file_path, 'r') as f:
        lines = f.readlines()
    
    data_rows = []
    for line in lines:
        line = line.strip()
        if line:
            parts = line.split('\t')
            # Ensure we have exactly 2 columns
            if len(parts) == 1:
                parts.append('')
            data_rows.append(parts[:2])
    
    return data_rows

def create_excel_for_iteration(iter_num, base_dir, metadata_dir, output_dir, data_type):
    """
    Create an Excel file for a single iteration.
    
    Args:
        iter_num: Iteration number
        base_dir: Base directory containing iteration folders
        metadata_dir: Directory containing metadata files
        output_dir: Directory to save Excel files
        data_type: 'gene' or 'isoform'
    """
    iter_dir = base_dir / f"iter_{iter_num}"
    deseq2_dir = iter_dir / f"deseq2_{data_type}s"
    
    if not deseq2_dir.exists():
        print(f"Warning: {deseq2_dir} not found, skipping iteration {iter_num}")
        return
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Tab 1: Metadata
    metadata_file = metadata_dir / f"metadata_trinity_iter_{iter_num}.txt"
    metadata_rows = read_metadata(metadata_file)
    
    ws_metadata = wb.create_sheet("Metadata", 0)
    # Add header
    ws_metadata.append(["Condition", "Sample", "R1", "R2"])
    # Make header bold
    for cell in ws_metadata[1]:
        cell.font = Font(bold=True)
    
    # Add metadata rows
    for row in metadata_rows:
        ws_metadata.append(row)
    
    # Tab 2: knockdown_PmSTAT_vs_knockdown_P
    comparison_file = deseq2_dir / f"knockdown_PmSTAT_vs_knockdown_P.P1e-3_C2.two_column_{data_type}s.txt"
    data = read_two_column_file(comparison_file)
    
    ws_comp1 = wb.create_sheet("knockdown_PmSTAT_vs_knockdown_P", 1)
    for row in data:
        ws_comp1.append(row)
    # Make first row bold (header)
    if ws_comp1.max_row > 0:
        for cell in ws_comp1[1]:
            cell.font = Font(bold=True)
    
    # Tab 3: WSSV_infection_vs_knockdown_PmS
    comparison_file = deseq2_dir / f"WSSV_infection_vs_knockdown_PmS.P1e-3_C2.two_column_{data_type}s.txt"
    data = read_two_column_file(comparison_file)
    
    ws_comp2 = wb.create_sheet("WSSV_infection_vs_knockdown_PmS", 2)
    for row in data:
        ws_comp2.append(row)
    # Make first row bold (header)
    if ws_comp2.max_row > 0:
        for cell in ws_comp2[1]:
            cell.font = Font(bold=True)
    
    # Tab 4: WSSV_infection_vs_knockdown_P-1
    comparison_file = deseq2_dir / f"WSSV_infection_vs_knockdown_P-1.P1e-3_C2.two_column_{data_type}s.txt"
    data = read_two_column_file(comparison_file)
    
    ws_comp3 = wb.create_sheet("WSSV_infection_vs_knockdown_P-1", 3)
    for row in data:
        ws_comp3.append(row)
    # Make first row bold (header)
    if ws_comp3.max_row > 0:
        for cell in ws_comp3[1]:
            cell.font = Font(bold=True)
    
    # Save Excel file
    output_file = output_dir / f"deseq2_{data_type}s_iter_{iter_num}.xlsx"
    wb.save(output_file)
    print(f"Created: {output_file}")

def main():
    """Main function to create Excel files for all iterations."""
    base_dir = Path("/home/luigui/Documents/2026/p_monodon_transcriptome/05_multiple_group_clustering/results_iterations")
    metadata_dir = Path("/home/luigui/Documents/2026/p_monodon_transcriptome/05_multiple_group_clustering/metadata_iterations")
    output_dir = Path("/home/luigui/Documents/2026/p_monodon_transcriptome/05_multiple_group_clustering/excel_outputs")
    
    # Create output directory if it doesn't exist
    output_dir.mkdir(exist_ok=True)
    
    if not base_dir.exists():
        print(f"Error: Base directory not found: {base_dir}")
        sys.exit(1)
    
    if not metadata_dir.exists():
        print(f"Error: Metadata directory not found: {metadata_dir}")
        sys.exit(1)
    
    # Process all iterations
    for i in range(1, 41):
        print(f"\nProcessing iteration {i}...")
        
        # Create Excel for genes
        create_excel_for_iteration(i, base_dir, metadata_dir, output_dir, "gene")
        
        # Create Excel for isoforms
        create_excel_for_iteration(i, base_dir, metadata_dir, output_dir, "isoform")
    
    print(f"\n✓ All Excel files created in: {output_dir}")
    print(f"  Total files: {len(list(output_dir.glob('*.xlsx')))}")

if __name__ == "__main__":
    main()
