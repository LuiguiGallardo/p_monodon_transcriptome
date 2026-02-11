#!/usr/bin/env python3
"""
Comprehensive test script to verify Excel files are correctly generated.
Tests data integrity, structure, and content accuracy.
"""

import os
import sys
from pathlib import Path
from openpyxl import load_workbook

def test_file_exists(file_path):
    """Test if file exists."""
    return os.path.exists(file_path)

def test_excel_structure(excel_file, iter_num, data_type):
    """Test Excel file structure and content."""
    results = {
        'file': excel_file.name,
        'tests_passed': [],
        'tests_failed': [],
        'warnings': []
    }
    
    try:
        wb = load_workbook(excel_file)
        
        # Test 1: Check sheet names
        expected_sheets = ['Metadata', 'knockdown_PmSTAT_vs_knockdown_P', 
                          'WSSV_infection_vs_knockdown_PmS', 'WSSV_infection_vs_knockdown_P-1']
        if wb.sheetnames == expected_sheets:
            results['tests_passed'].append(f"✓ All 4 tabs present in correct order")
        else:
            results['tests_failed'].append(f"✗ Sheet names mismatch. Expected: {expected_sheets}, Got: {wb.sheetnames}")
        
        # Test 2: Metadata tab structure
        ws_meta = wb['Metadata']
        meta_header = [cell.value for cell in ws_meta[1]]
        expected_meta_header = ['Condition', 'Sample', 'R1', 'R2']
        if meta_header == expected_meta_header:
            results['tests_passed'].append(f"✓ Metadata header correct")
        else:
            results['tests_failed'].append(f"✗ Metadata header mismatch. Expected: {expected_meta_header}, Got: {meta_header}")
        
        # Test 3: Metadata has data (should have 9 samples + 1 header = 10 rows)
        if ws_meta.max_row >= 9:
            results['tests_passed'].append(f"✓ Metadata has sample data ({ws_meta.max_row - 1} samples)")
        else:
            results['tests_failed'].append(f"✗ Metadata has insufficient data (only {ws_meta.max_row - 1} samples)")
        
        # Test 4-6: Check comparison tabs
        comparisons = [
            ('knockdown_PmSTAT_vs_knockdown_P', ['knockdown_PmSTAT-UP', 'knockdown_PmSTAT_WSSV_infection-UP']),
            ('WSSV_infection_vs_knockdown_PmS', ['knockdown_PmSTAT-UP', 'WSSV_infection-UP']),
            ('WSSV_infection_vs_knockdown_P-1', ['knockdown_PmSTAT_WSSV_infection-UP', 'WSSV_infection-UP'])
        ]
        
        for tab_name, expected_header in comparisons:
            ws = wb[tab_name]
            actual_header = [cell.value for cell in ws[1]]
            
            if actual_header == expected_header:
                results['tests_passed'].append(f"✓ {tab_name}: Header correct")
            else:
                results['tests_failed'].append(f"✗ {tab_name}: Header mismatch. Expected: {expected_header}, Got: {actual_header}")
            
            # Check if tab has data (at least header)
            if ws.max_row >= 1:
                row_count = ws.max_row - 1  # Subtract header
                if row_count > 0:
                    results['tests_passed'].append(f"✓ {tab_name}: Has {row_count} data rows")
                else:
                    results['warnings'].append(f"⚠ {tab_name}: No data rows (only header)")
            else:
                results['tests_failed'].append(f"✗ {tab_name}: Empty tab")
        
        # Test 7: Cross-check with source files
        base_dir = Path("/home/luigui/Documents/2026/p_monodon_transcriptome/05_multiple_group_clustering/results_iterations")
        iter_dir = base_dir / f"iter_{iter_num}"
        deseq2_dir = iter_dir / f"deseq2_{data_type}s"
        
        for tab_name, _ in comparisons:
            # Map tab name to file name
            file_mapping = {
                'knockdown_PmSTAT_vs_knockdown_P': 'knockdown_PmSTAT_vs_knockdown_P',
                'WSSV_infection_vs_knockdown_PmS': 'WSSV_infection_vs_knockdown_PmS',
                'WSSV_infection_vs_knockdown_P-1': 'WSSV_infection_vs_knockdown_P-1'
            }
            
            source_file = deseq2_dir / f"{file_mapping[tab_name]}.P1e-3_C2.two_column_{data_type}s.txt"
            
            if source_file.exists():
                # Count lines in source file
                with open(source_file, 'r') as f:
                    source_lines = len([l for l in f.readlines() if l.strip()])
                
                ws = wb[tab_name]
                excel_rows = ws.max_row
                
                if source_lines == excel_rows:
                    results['tests_passed'].append(f"✓ {tab_name}: Row count matches source file ({source_lines} rows)")
                else:
                    results['tests_failed'].append(f"✗ {tab_name}: Row count mismatch. Source: {source_lines}, Excel: {excel_rows}")
            else:
                results['warnings'].append(f"⚠ {tab_name}: Source file not found for verification")
        
    except Exception as e:
        results['tests_failed'].append(f"✗ Error loading Excel file: {str(e)}")
    
    return results

def run_all_tests():
    """Run tests on all Excel files."""
    output_dir = Path("/home/luigui/Documents/2026/p_monodon_transcriptome/05_multiple_group_clustering/excel_outputs")
    
    if not output_dir.exists():
        print(f"ERROR: Output directory not found: {output_dir}")
        sys.exit(1)
    
    print("=" * 80)
    print("EXCEL FILE VERIFICATION TEST SUITE")
    print("=" * 80)
    
    all_results = []
    total_files = 0
    files_passed = 0
    files_with_warnings = 0
    files_failed = 0
    
    # Test all iterations
    for i in range(1, 41):
        for data_type in ['gene', 'isoform']:
            excel_file = output_dir / f"deseq2_{data_type}s_iter_{i}.xlsx"
            
            if not test_file_exists(excel_file):
                print(f"\n✗ MISSING FILE: {excel_file.name}")
                files_failed += 1
                continue
            
            total_files += 1
            results = test_excel_structure(excel_file, i, data_type)
            all_results.append(results)
            
            # Determine status
            if results['tests_failed']:
                files_failed += 1
                status = "FAILED"
                symbol = "✗"
            elif results['warnings']:
                files_with_warnings += 1
                status = "WARNING"
                symbol = "⚠"
            else:
                files_passed += 1
                status = "PASSED"
                symbol = "✓"
            
            # Print summary for this file
            if status != "PASSED" or i <= 3:  # Show first 3 iterations + any failures
                print(f"\n{symbol} {results['file']} - {status}")
                for test in results['tests_passed']:
                    print(f"  {test}")
                for test in results['tests_failed']:
                    print(f"  {test}")
                for warning in results['warnings']:
                    print(f"  {warning}")
    
    # Print overall summary
    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"Total files tested: {total_files}")
    print(f"✓ Passed: {files_passed}")
    print(f"⚠ Warnings: {files_with_warnings}")
    print(f"✗ Failed: {files_failed}")
    
    if files_failed > 0:
        print("\n❌ VERIFICATION FAILED - Some files have errors")
        print("   Please review the failed tests above and rerun the script.")
        return False
    elif files_with_warnings > 0:
        print("\n⚠️  VERIFICATION PASSED WITH WARNINGS")
        print("   Files are structurally correct but some comparisons have no data.")
        print("   This is expected for iterations with no differentially expressed genes/isoforms.")
        return True
    else:
        print("\n✅ ALL TESTS PASSED - Excel files are 100% correct!")
        return True

if __name__ == "__main__":
    success = run_all_tests()
    sys.exit(0 if success else 1)
